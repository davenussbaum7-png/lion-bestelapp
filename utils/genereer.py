"""
Genereer piklijsten en paklijsten vanuit web-bestellingen.
Gebaseerd op verwerk.py logica, aangepast voor webapp-gebruik.
"""
import io
import re
import zipfile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
# ─── Opmaak ───────────────────────────────────────────────────────────────────
KLEUR_KOPTEKST   = "1C1C1C"
KLEUR_SECTIE_VUL = "C0C0C0"
KLEUR_WIT        = "FFFFFF"
KLEUR_ZEBRA      = "F5F5F5"
KLEUR_TOTAAL     = "404040"
DATA_SIZE = 12
FONT_KOPTEKST   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_SECTIE     = Font(name="Calibri", bold=True, color="000000", size=12)
FONT_NORMAAL    = Font(name="Calibri", size=DATA_SIZE)
FONT_ITALIC     = Font(name="Calibri", italic=True, size=DATA_SIZE)
FONT_NIET_VOORR = Font(name="Calibri", bold=True, size=DATA_SIZE)
RAND_DUN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
RAND_SECTIE = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)
RAND_NIET_VOORR = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)
# ─── Verrijken en sorteren ────────────────────────────────────────────────────
def bouw_artikellijst(winkelnaam, orders, dbo_orders, sap_data, artikelen_db):
    """
    Combineer winkelbestelling + SAP-data tot een gesorteerde artikellijst.
    orders      : {ean: quantity}  — winkelbestelling
    dbo_orders  : [{sectie, artikel, quantity}]
    sap_data    : {ean: {stuks_verkocht, voorraad_centraal, artikel}}
    artikelen_db: [{ean, artikel, sectie, pad_code, volgorde}]
    """
    # EAN → catalogus-info
    catalogus = {a["ean"]: a for a in artikelen_db if a.get("ean")}
    # Sectie → pad_code mapping (voor DBO en SAP-items zonder EAN in catalogus)
    sectie_pad = {}
    for a in artikelen_db:
        s = a.get("sectie")
        p = a.get("pad_code")
        if s and p and s not in sectie_pad:
            sectie_pad[s] = p

    def normalize_pad_code(raw):
        """Normaliseer pad_code naar canonical vorm.
        '01 1 PERS.DBO' → '1', '03 3 Pers.DBO' → '3', '3A' → '3A', 'MATRASSEN' → ''
        """
        if not raw:
            return ""
        raw = str(raw).strip()
        # Al correct: "1", "3A", "13A", "15a"
        if re.match(r'^\d+[A-Za-z]*$', raw):
            return raw
        # Sectienaam met 1-2 cijfer prefix gevolgd door spatie: "01 1 PERS.DBO" → "1"
        m = re.match(r'^(\d{1,2})\s', raw)
        if m:
            return str(int(m.group(1)))
        return ""

    bestelde_eans = set(orders.keys())
    resultaat = []
    # 1. Winkelbestellingen
    for ean, besteld in orders.items():
        if besteld <= 0:
            continue
        cat = catalogus.get(ean, {})
        sap = sap_data.get(ean, {})
        stuks = sap.get("stuks_verkocht", 0) or 0
        voorraad = sap.get("voorraad_centraal", 999)
        if voorraad is None:
            voorraad = 999
        resultaat.append({
            "type":       "ARTIKEL",
            "ean":        ean,
            "artikel":    cat.get("artikel") or sap.get("artikel") or ean,
            "sectie":     cat.get("sectie") or sap.get("groepsnaam") or "",
            "pad_code":   normalize_pad_code(cat.get("pad_code") or ""),
            "volgorde":   cat.get("volgorde") or 9999,
            "besteld":    besteld,
            "sap":        stuks,
            "op_voorraad": voorraad > 0,
        })
    # 2. SAP-only (niet besteld door winkel, wel in SAP)
    for ean, sap in sap_data.items():
        if ean in bestelde_eans:
            continue
        stuks = sap.get("stuks_verkocht", 0) or 0
        if stuks <= 0:
            continue
        cat = catalogus.get(ean, {})
        voorraad = sap.get("voorraad_centraal", 999)
        if voorraad is None:
            voorraad = 999
        sectie_sap = cat.get("sectie") or sap.get("groepsnaam") or ""
        raw_pad_sap = cat.get("pad_code") or sectie_pad.get(sectie_sap) or sectie_sap
        resultaat.append({
            "type":       "SAP",
            "ean":        ean,
            "artikel":    cat.get("artikel") or sap.get("artikel") or ean,
            "sectie":     sectie_sap,
            "pad_code":   normalize_pad_code(raw_pad_sap),
            "volgorde":   cat.get("volgorde") or 9999,
            "besteld":    0,
            "sap":        stuks,
            "op_voorraad": voorraad > 0,
        })
    # 3. DBO vrije regels
    for dbo in dbo_orders:
        qty = dbo.get("quantity", 0) or 0
        if qty <= 0:
            continue
        sectie_dbo = dbo.get("sectie", "DBO")
        # Gebruik sectie_pad lookup, dan sectienaam zelf als fallback voor pad-afleiding
        raw_pad_dbo = sectie_pad.get(sectie_dbo) or sectie_dbo
        resultaat.append({
            "type":       "DBO",
            "ean":        None,
            "artikel":    dbo.get("artikel", ""),
            "sectie":     sectie_dbo,
            "pad_code":   normalize_pad_code(raw_pad_dbo),
            "volgorde":   -1,  # DBO secties 01-04 bovenaan
            "besteld":    qty,
            "sap":        0,
            "op_voorraad": True,
        })
    # Sorteren: pad_code numerisch (1, 2, 2A, 3, 12 ipv 1, 12, 2, 3...)
    def pad_sort_key(code):
        if not code:
            return (9999, "")
        m = re.match(r'^(\d+)([A-Za-z]*)$', str(code).strip())
        if m:
            return (int(m.group(1)), m.group(2).upper())
        return (9999, str(code))

    resultaat.sort(key=lambda a: (pad_sort_key(a["pad_code"]), a["volgorde"], a["artikel"]))
    return resultaat
# ─── Piklijst schrijven ───────────────────────────────────────────────────────
def schrijf_piklijst(winkelnaam, artikelen):
    """Geeft bytes van de piklijst xlsx terug."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Piklijst"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = "1:3"
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top  = ws.page_margins.bottom = 0.5
    for i, breedte in enumerate([5, 28, 57, 8, 8, 8, 8, 14, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = breedte
    # Titel
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"PIKLIJST — {winkelnaam.upper()}"
    c.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=KLEUR_KOPTEKST)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    # Legenda
    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = "VET + dikke rand = NIET OP VOORRAAD   |   Cursief = SAP aanvulling of DBO vrije invoer"
    c.font  = Font(name="Calibri", size=8, italic=True, color="444444")
    c.fill  = PatternFill("solid", fgColor="F0F0F0")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 14
    # Kopteksten
    headers = ["Pad", "Sectie", "Artikel", "Besteld", "SAP", "Totaal", "Gepakt □", "EAN", "Opmerking"]
    for ci, h in enumerate(headers, 1):
        cel = ws.cell(row=3, column=ci, value=h)
        cel.font      = FONT_KOPTEKST
        cel.fill      = PatternFill("solid", fgColor=KLEUR_KOPTEKST)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border    = RAND_DUN
    ws.row_dimensions[3].height = 18
    huidige_sectie = None
    huidige_pad    = None
    rij_nr         = 4
    for art in artikelen:
        sectie   = art.get("sectie") or ""
        pad_code = art.get("pad_code") or ""
        # Nieuwe pagina bij wisseling van pad_code (alleen als pad_code ingevuld is)
        if pad_code and huidige_pad is not None and pad_code != huidige_pad:
            ws.row_breaks.append(Break(id=rij_nr - 1))
        huidige_pad = pad_code
        if sectie != huidige_sectie:
            huidige_sectie = sectie
            ws.merge_cells(f"A{rij_nr}:I{rij_nr}")
            sc = ws.cell(row=rij_nr, column=1, value=f"  {sectie}")
            sc.font      = FONT_SECTIE
            sc.fill      = PatternFill("solid", fgColor=KLEUR_SECTIE_VUL)
            sc.alignment = Alignment(vertical="center", indent=1)
            sc.border    = RAND_SECTIE
            ws.row_dimensions[rij_nr].height = 16
            rij_nr += 1
        niet_op_voorraad = not art["op_voorraad"] and art["sap"] > 0
        is_dbo           = art["type"] == "DBO"
        is_sap_only      = art["type"] == "SAP" and art["besteld"] == 0
        if niet_op_voorraad:
            font_data = FONT_NIET_VOORR
            opmerking = "!!! NIET OP VOORRAAD"
            rand      = RAND_NIET_VOORR
        elif is_dbo or is_sap_only:
            font_data = FONT_ITALIC
            opmerking = "DBO — vrije invoer winkel" if is_dbo else "SAP aanvulling"
            rand      = RAND_DUN
        else:
            font_data = FONT_NORMAAL
            opmerking = ""
            rand      = RAND_DUN
        totaal = (art["besteld"] or 0) + (art["sap"] or 0)
        waarden = [
            pad_code, sectie, art["artikel"],
            art["besteld"] or "",
            art["sap"] if art["sap"] else "",
            totaal if totaal else "",
            "", art["ean"] or "", opmerking,
        ]
        for ci, waarde in enumerate(waarden, 1):
            cel = ws.cell(row=rij_nr, column=ci, value=waarde)
            cel.font      = font_data
            cel.fill      = PatternFill("solid", fgColor=KLEUR_WIT)
            cel.border    = rand
            cel.alignment = Alignment(
                horizontal="center" if ci in (1, 4, 5, 6, 7) else "left",
                vertical="center",
                indent=1 if ci in (2, 3, 9) else 0,
            )
        ws.row_dimensions[rij_nr].height = 18
        rij_nr += 1
    # Totaalregel
    totaal_stuks    = sum((a["besteld"] or 0) + (a["sap"] or 0) for a in artikelen)
    niet_voorraad_n = sum(1 for a in artikelen if not a["op_voorraad"] and a["sap"] > 0)
    ws.merge_cells(f"A{rij_nr}:I{rij_nr}")
    tc = ws.cell(row=rij_nr, column=1,
                 value=f"  TOTAAL: {totaal_stuks} stuks  |  {len(artikelen)} regels  |  {niet_voorraad_n} NIET OP VOORRAAD")
    tc.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=KLEUR_TOTAAL)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[rij_nr].height = 18
    ws.freeze_panes = "A4"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
# ─── Paklijst schrijven ───────────────────────────────────────────────────────
def schrijf_paklijst(winkelnaam, piklijst_bytes):
    """
    Leest gecorrigeerde piklijst (bytes) en genereert paklijst.
    Kolom 6 (Totaal) is leidend — Wouter past dat aan bij NIET OP VOORRAAD.
    """
    buf_in = io.BytesIO(piklijst_bytes)
    wb_in  = openpyxl.load_workbook(buf_in, data_only=True)
    ws_in  = wb_in.active
    artikelen = []
    for rij in ws_in.iter_rows(min_row=4, values_only=True):
        pad, sectie, artikel = rij[0], rij[1], rij[2]
        totaal = rij[5]
        ean    = rij[7] if len(rij) > 7 else None
        if sectie is None and artikel is None:
            continue
        if pad and str(pad).strip().startswith("TOTAAL"):
            continue
        if artikel is None:
            continue
        try:
            totaal_val = int(float(str(totaal))) if totaal is not None else 0
        except (ValueError, TypeError):
            totaal_val = 0
        if totaal_val <= 0:
            continue
        artikelen.append({
            "ean":     str(ean).strip() if ean else "",
            "artikel": str(artikel).strip(),
            "totaal":  totaal_val,
        })
    # Paklijst bouwen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paklijst"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = "1:2"
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top  = ws.page_margins.bottom = 0.7
    for ci, breedte in zip(range(1, 4), [18, 55, 10]):
        ws.column_dimensions[get_column_letter(ci)].width = breedte
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = f"PAKLIJST — {winkelnaam.upper()}"
    t.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor=KLEUR_KOPTEKST)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(["EAN", "Omschrijving", "Aantal"], 1):
        cel = ws.cell(row=2, column=ci, value=h)
        cel.font      = FONT_KOPTEKST
        cel.fill      = PatternFill("solid", fgColor=KLEUR_KOPTEKST)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border    = RAND_DUN
    ws.row_dimensions[2].height = 18
    def pak_sort(a):
        try:
            return (0, int(a["ean"]), a["artikel"])
        except (ValueError, TypeError):
            return (1, 0, a["artikel"])
    gesorteerd = sorted(artikelen, key=pak_sort)
    rij_nr = 3
    zebra  = False
    for art in gesorteerd:
        achtergrond = KLEUR_ZEBRA if zebra else KLEUR_WIT
        for ci, waarde in enumerate([art["ean"] or "—", art["artikel"], art["totaal"]], 1):
            cel = ws.cell(row=rij_nr, column=ci, value=waarde)
            cel.font      = FONT_NORMAAL
            cel.fill      = PatternFill("solid", fgColor=achtergrond)
            cel.border    = RAND_DUN
            cel.alignment = Alignment(
                horizontal="center" if ci in (1, 3) else "left",
                vertical="center",
                indent=1 if ci == 2 else 0,
            )
        ws.row_dimensions[rij_nr].height = 15
        rij_nr += 1
        zebra = not zebra
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
# ─── SAP xlsx inlezen ────────────────────────────────────────────────────────
def lees_sap_xlsx(bestand_bytes) -> tuple:
    """
    Leest een SAP-export xlsx (bytes).
    Geeft (winkelnaam, [{ean, artikel, stuks_verkocht, voorraad_centraal}]) terug.
    """
    buf = io.BytesIO(bestand_bytes)
    wb  = openpyxl.load_workbook(buf, data_only=True)
    sheetnaam = "Blad1" if "Blad1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheetnaam]
    headers = {}
    for rij in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if any(v is not None for v in rij):
            for ci, h in enumerate(rij):
                if h:
                    headers[str(h).strip().lower()] = ci
            break
    def kolom(opties):
        for n in opties:
            if n in headers:
                return headers[n]
        return None
    ci_winkel   = kolom(["magazijnnaam"])
    ci_ean      = kolom(["artikelnummer", "artikel_nummer"])
    ci_artikel  = kolom(["dscription", "description", "omschrijving"])
    ci_verkocht = kolom(["stuks_verkocht", "stuks verkocht"])
    ci_voorraad = kolom(["voorraad_centraal", "voorraad centraal"])
    ci_groep    = kolom(["groepsnaam"])
    if ci_ean is None or ci_verkocht is None:
        return None, []
    sap_regels = []
    winkelnaam = None
    for rij in ws.iter_rows(min_row=2, values_only=True):
        if not any(rij):
            continue
        if ci_winkel is not None and rij[ci_winkel] and winkelnaam is None:
            winkelnaam = str(rij[ci_winkel]).strip()
        ean_raw  = rij[ci_ean]
        verkocht = rij[ci_verkocht]
        if ean_raw is None or verkocht is None:
            continue
        try:
            ean   = str(int(float(str(ean_raw))))
            stuks = int(float(str(verkocht)))
        except (ValueError, TypeError):
            continue
        if stuks <= 0:
            continue
        voorraad = 999
        if ci_voorraad is not None and rij[ci_voorraad] is not None:
            try:
                voorraad = int(float(str(rij[ci_voorraad])))
            except (ValueError, TypeError):
                pass
        artikel = ""
        if ci_artikel is not None and rij[ci_artikel]:
            artikel = str(rij[ci_artikel]).strip()
        groep = ""
        if ci_groep is not None and rij[ci_groep]:
            groep = str(rij[ci_groep]).strip()
        sap_regels.append({
            "ean":               ean,
            "artikel":           artikel,
            "stuks_verkocht":    stuks,
            "voorraad_centraal": voorraad,
            "groepsnaam":        groep,
        })
    return winkelnaam, sap_regels


# ─── Padcodes xlsx inlezen ────────────────────────────────────────────────────
def lees_padcodes_xlsx(bestand_bytes) -> dict:
    """
    Leest een Excel met minimaal kolommen EAN en Pad_code.
    Geeft {ean: pad_code} terug.
    Herkent ook: 'Padcode', 'Pad code', 'pad', 'pad_nr'.
    """
    buf = io.BytesIO(bestand_bytes)
    wb  = openpyxl.load_workbook(buf, data_only=True)
    ws  = wb.active

    headers = {}
    for rij in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if any(v is not None for v in rij):
            for ci, h in enumerate(rij):
                if h:
                    headers[str(h).strip().lower().replace(" ", "_")] = ci
            break

    def kolom(opties):
        for n in opties:
            if n in headers:
                return headers[n]
        return None

    ci_ean = kolom(["ean", "artikelnummer", "barcode", "artikel_nummer"])
    ci_pad = kolom(["pad_code", "padcode", "pad", "pad_nr", "pad_nummer"])

    if ci_ean is None or ci_pad is None:
        return {}

    pad_codes = {}
    for rij in ws.iter_rows(min_row=2, values_only=True):
        if not any(rij):
            continue
        ean_raw = rij[ci_ean]
        pad_raw = rij[ci_pad]
        if ean_raw is None or pad_raw is None:
            continue
        try:
            ean = str(int(float(str(ean_raw))))
        except (ValueError, TypeError):
            ean = str(ean_raw).strip()
        pad_codes[ean] = str(pad_raw).strip()

    return pad_codes
