"""
Eenmalig setup-script: laad artikelcatalogus + winkels in Supabase.

Gebruik:
  1. Zet je Supabase URL en key in .streamlit/secrets.toml
  2. Kopieer Winkel_Invoer_voor_winkels_v15.xlsx naar dezelfde map als dit script
  3. Pas WINKELS aan (namen + PIN-codes)
  4. Voer uit: python setup/laad_catalogus.py

Dit script MAG je meerdere keren draaien — het verwijdert eerst de oude data.
"""

import sys
import os
import openpyxl
from supabase import create_client

# ─── Configuratie ─────────────────────────────────────────────────────────────
# Pas dit aan:
SUPABASE_URL = "https://JOUW-PROJECT.supabase.co"
SUPABASE_KEY = "JOUW-ANON-KEY"

CATALOGUS_BESTAND = "Winkel_Invoer_voor_winkels_v15.xlsx"

# Winkelnamen moeten EXACT overeenkomen met cel I1 in de WinkelInvoer-formulieren
# en met de Magazijnnaam in de SAP-exports.
WINKELS = [
    {"name": "Delft",        "pin": "1234"},
    {"name": "Rotterdam",    "pin": "2345"},
    {"name": "Den Haag",     "pin": "3456"},
    {"name": "Utrecht",      "pin": "4567"},
    {"name": "Amsterdam",    "pin": "5678"},
    {"name": "Leiden",       "pin": "6789"},
    {"name": "Zoetermeer",   "pin": "7890"},
    {"name": "Naaldwijk",    "pin": "8901"},
    {"name": "Rijswijk",     "pin": "9012"},
    {"name": "Wateringen",   "pin": "0123"},
]

# ─── Verbinding ───────────────────────────────────────────────────────────────
db = create_client(SUPABASE_URL, SUPABASE_KEY)


def laad_catalogus():
    print(f"Catalogus lezen uit: {CATALOGUS_BESTAND}")
    wb = openpyxl.load_workbook(CATALOGUS_BESTAND, data_only=True)

    # Secties op volgorde uit dropdown-kolom H
    ws_cat = wb["Artikelcatalogus"]
    secties_volgorde = {}
    for rij in ws_cat.iter_rows(min_row=2, values_only=True):
        if len(rij) > 7 and rij[7]:
            s = str(rij[7]).strip()
            if s not in secties_volgorde:
                secties_volgorde[s] = len(secties_volgorde)

    # DBO-secties 01-04 bovenaan (volgorde negatief)
    dbo_secties = ["01 1 PERS.DBO", "02 2 PERS.DBO", "03 3 Pers.DBO", "04 260 BR.DBO"]
    for i, s in enumerate(dbo_secties):
        secties_volgorde[s] = -(len(dbo_secties) - i)

    # Artikelen lezen
    artikelen = []
    volgorde_counter = {}

    for rij in ws_cat.iter_rows(min_row=2, values_only=True):
        sectie, artikel, ean, markering = rij[0], rij[1], rij[2], rij[3] if len(rij) > 3 else None
        if markering == "KOP" or not artikel:
            continue
        if not sectie:
            continue

        sectie_str = str(sectie).strip()
        artikel_str = str(artikel).strip()

        if ean is not None:
            try:
                ean_str = str(int(float(str(ean))))
            except (ValueError, OverflowError):
                ean_str = str(ean).strip()
        else:
            ean_str = None

        sectie_idx = secties_volgorde.get(sectie_str, 9999)
        volgorde_counter[sectie_str] = volgorde_counter.get(sectie_str, 0) + 1
        volgorde = sectie_idx * 10000 + volgorde_counter[sectie_str]

        artikelen.append({
            "ean":      ean_str,
            "artikel":  artikel_str,
            "sectie":   sectie_str,
            "pad_code": "",           # later te vullen vanuit Bestellijst_Geconsolideerd
            "volgorde": volgorde,
        })

    print(f"  {len(artikelen)} artikelen gevonden in {len(secties_volgorde)} secties")

    # Opslaan in Supabase
    print("  Oude artikelen verwijderen...")
    db.table("articles").delete().neq("id", 0).execute()

    print("  Nieuwe artikelen opslaan...")
    batch_size = 200
    for i in range(0, len(artikelen), batch_size):
        batch = artikelen[i:i + batch_size]
        db.table("articles").insert(batch).execute()
        print(f"    {min(i + batch_size, len(artikelen))}/{len(artikelen)} opgeslagen")

    print(f"  ✅ Catalogus geladen: {len(artikelen)} artikelen")


def laad_winkels():
    print("Winkels opslaan...")
    db.table("stores").delete().neq("id", 0).execute()
    db.table("stores").insert(WINKELS).execute()
    print(f"  ✅ {len(WINKELS)} winkels opgeslagen")


if __name__ == "__main__":
    laad_catalogus()
    laad_winkels()
    print("\nSetup klaar! Je kunt de app nu starten met: streamlit run app.py")
